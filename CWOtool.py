import os
import random
import sys
import xlwings as xw
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from openpyxl import load_workbook, Workbook
from ui_Main import Ui_cwo_generator_tool
from PyQt5 import QtWidgets
import pandas as pd


class appWindow(QtWidgets.QMainWindow, Ui_cwo_generator_tool):
    def __init__(self, parent=None):
        super(appWindow, self).__init__(parent)
        self.setupUi(self)
        self.setWindowIcon(QIcon('icon.ico'))

        self.config_2G_initialGUI()
        self.bscName.currentTextChanged.connect(lambda: [self.calculate_bscID(), self.set_bcsuID()])
        self.sectors_SRAN.valueChanged.connect(
            lambda: [self.sector_change(), self.set_2G_config(), self.CI_TRX_change(), self.validate_inputs()])
        self.tech_2G.currentTextChanged.connect(
            lambda: [self.sector_change(), self.set_2G_config(), self.CI_TRX_change(), self.validate_inputs()])
        self.bcfID.textChanged.connect(lambda: [self.sector_change(), self.set_2G_config()])
        self.cwo2G.setEnabled(False)
        self.cwo3G.setEnabled(False)
        self.cwo4G.setEnabled(False)
        self.cwo2G.clicked.connect(self.create_2G_dataframe)
        self.cwo3G.clicked.connect(self.create_3G_dataframe)
        self.cwo4G.clicked.connect(self.create_4G_dataframe)
        self.upload_Ucells.clicked.connect(self.load_Ucells_data)
        self.upload_Lcells.clicked.connect(self.load_Lcells_data)
        self.cellU_df = pd.DataFrame()
        self.cellL_df = pd.DataFrame()
        self.validateData.clicked.connect(self.validate_inputs)

    # Getters for common site data
    def get_siteName(self):
        return self.siteName_SRAN.text()

    def get_siteCode(self):
        return self.siteCode_SRAN.text()

    def get_site_area(self):
        area = self.get_siteCode()[-2:]
        return area

    def get_mngtIP(self):
        return self.mngtIP.text()

    def get_mngtGW(self):
        return self.mngtGW.text()

    # Getters for 3G site data
    def load_Ucells_data(self):
        fileName = QFileDialog.getOpenFileName(filter="Excel (*.xlsx *.xls *.csv)")
        if fileName[1] == "":
            QMessageBox.about(self, 'Caution', "No file uploaded")
        else:
            path = fileName[0]
            cellData3G = pd.read_excel(path, sheet_name=0)
            cellData3G_header = list(cellData3G.columns.values)
            for i in cellData3G_header:
                self.cellID_column.addItem(i)
            for i in cellData3G_header:
                self.cz_column.addItem(i)
            for i in cellData3G_header:
                self.lacU_column.addItem(i)

            self.cellU_df = cellData3G

            QMessageBox.about(self, 'Done', "3G Cells data Uploaded successfully")
            return self.cellU_df

    def get_wbtsID(self):
        return self.wbtsID.text()

    def get_rncID(self):
        return self.rncID.text()

    def get_rncCard(self):
        return self.rncCard.text()

    def get_portSCTP(self):
        return self.portSCTP.text()

    def get_voiceCommittedBW(self):
        return self.vCommittedBW.text()

    def get_voiceRouteBW(self):
        return self.vRouteBW.text()

    def get_defaultCommittedBW(self):
        return self.dCommittedBW.text()

    def get_defaultRouteBW(self):
        return self.dRouteBW.text()

    def get_routeBW(self):
        return self.routeBW.text()

    # Getters for LTE configuration
    def get_secIP(self):
        return self.secIP.text()

    def get_secGW(self):
        return self.secGW.text()

    def get_coreSite(self):
        return self.coreSite.text()

    def get_serverCA(self):
        return self.serverCA.text()

    def load_Lcells_data(self):
        fileName = QFileDialog.getOpenFileName(filter="Excel (*.xlsx *.xls *.csv)")
        if fileName[1] == "":
            QMessageBox.about(self, 'Caution', "No file uploaded")
        else:
            path = fileName[0]
            cellDataLTE = pd.read_excel(path, sheet_name=0)
            cellDataLTE_header = list(cellDataLTE.columns.values)

            if 'Unnamed' in str(cellDataLTE_header):
                cellDataLTE.columns = cellDataLTE.iloc[0]
                cellDataLTE = cellDataLTE.iloc[1:]
                cellDataLTE_header = list(cellDataLTE.columns.values)
            else:
                pass

            for i in cellDataLTE_header:
                self.cellID_LTE.addItem(i)
            for i in cellDataLTE_header:
                self.cz_LTE_column.addItem(i)
            for i in cellDataLTE_header:
                self.localCI_LTE.addItem(i)
            for i in cellDataLTE_header:
                self.cellName_LTE.addItem(i)
            for i in cellDataLTE_header:
                self.eNodeB_LTE.addItem(i)

            self.cellL_df = cellDataLTE
            QMessageBox.about(self, 'Done', "4G Cells data Uploaded successfully")
            return self.cellL_df

    # Getters for 2G configuration

    def get_bcfID(self):
        return self.bcfID.text()

    def calculate_bscID(self):
        bsc_dict = {'362965': 'MINYA01', '391778': 'SHARM01', '392853': 'SINAI', '402772': 'SOHAG02',
                    '403736': 'ASWAN02', '403737': 'BNISUEF02', '403738': 'BENISUEF01', '406889': 'ASYOUT01',
                    '406893': 'ASYOUT02', '406894': 'QENA01', '406895': 'QENA02', '410223': 'MINYA',
                    '891392': 'ASWAN01', '337193': 'MANSHIA01', '359891': 'FAQOUS01', '359892': 'ZAGAZIG01',
                    '360143': 'NEWAWAYED02', '362962': 'KFDW01', '373657': 'TANTA01', '373658': 'HIHYA01',
                    '373661': 'MINUF01', '376491': 'MIAMI01', '387594': 'ZAGAZIG02', '387596': 'DAMAS01',
                    '387597': 'KFRELSHKH01', '387667': 'NEWAWAYED01', '392835': 'MANSOURA01', '392854': 'BANHA01',
                    '395856': 'SEMOUHA02', '396623': 'SMOUHA01', '397646': 'AMERYA01', '401358': 'SHIBIN01',
                    '402792': 'Banha mcBSC', '529518': 'PortSaid', '881026': 'DUMYAT01', '891393': 'MANSHIA02',
                    '912852': 'DAMANHOUR01'}

        for k, v in bsc_dict.items():
            if v == self.get_bscName():
                return self.set_bscID(k)

    def get_bscID(self):
        return self.bscID.text()

    def get_bscName(self):
        return self.bscName.currentText()

    def set_bscID(self, x):
        self.bscID.setText(x)

    def set_bcsuID(self):
        x = random.randint(0, 6)
        return self.bcsuID.setText(str(x))

    def get_bcsuID(self):
        return self.bcsuID.text()

    def get_lacG(self):
        return self.lacGSM.text()

    def get_racG(self):
        return self.racG.text()

    def get_cuPlaneIP(self):
        return self.cuIP.text()

    def get_usedETP(self):
        return self.usedETP.text()

    def get_num_sectors(self):
        return self.sectors_SRAN.value()

    def get_tech_2G(self):
        return self.tech_2G.currentText()

    def config_2G_initialGUI(self):
        input_list_bool = [False, False, False, False, False, False, False, False, False, False, False, False, False,
                           False, False, False, False, False, False, False, False, False, False, False, False, False,
                           False, False, False, False, False, False]
        input_list = [self.bcf_gS1.setEnabled(input_list_bool[0]), self.bts_gS1.setEnabled(input_list_bool[1]),
                      self.gS1_CI.setEnabled(input_list_bool[2]), self.gS1_TRX.setEnabled(input_list_bool[3]),
                      self.bcf_gS2.setEnabled(input_list_bool[4]), self.bts_gS2.setEnabled(input_list_bool[5]),
                      self.gS2_CI.setEnabled(input_list_bool[6]), self.gS2_TRX.setEnabled(input_list_bool[7]),
                      self.bcf_gS3.setEnabled(input_list_bool[8]), self.bts_gS3.setEnabled(input_list_bool[9]),
                      self.gS3_CI.setEnabled(input_list_bool[10]), self.gS3_TRX.setEnabled(input_list_bool[11]),
                      self.bcf_gS4.setEnabled(input_list_bool[12]), self.bts_gS4.setEnabled(input_list_bool[13]),
                      self.gS4_CI.setEnabled(input_list_bool[14]), self.gS4_TRX.setEnabled(input_list_bool[15]),
                      self.bcf_dS1.setEnabled(input_list_bool[16]), self.bts_dS1.setEnabled(input_list_bool[17]),
                      self.dS1_CI.setEnabled(input_list_bool[18]), self.dS1_TRX.setEnabled(input_list_bool[19]),
                      self.bcf_dS2.setEnabled(input_list_bool[20]), self.bts_dS2.setEnabled(input_list_bool[21]),
                      self.dS2_CI.setEnabled(input_list_bool[22]), self.dS2_TRX.setEnabled(input_list_bool[23]),
                      self.bcf_dS3.setEnabled(input_list_bool[24]), self.bts_dS3.setEnabled(input_list_bool[25]),
                      self.dS3_CI.setEnabled(input_list_bool[26]), self.dS3_TRX.setEnabled(input_list_bool[27]),
                      self.bcf_dS4.setEnabled(input_list_bool[28]), self.bts_dS4.setEnabled(input_list_bool[29]),
                      self.dS4_CI.setEnabled(input_list_bool[30]), self.dS4_TRX.setEnabled(input_list_bool[31])]
        return input_list

    def sector_change(self):
        input_list_bool = [False, False, False, False, False, False, False, False, False, False, False, False, False,
                           False, False, False, False, False, False, False, False, False, False, False, False, False,
                           False, False, False, False, False, False]
        x = self.get_num_sectors()
        y = self.get_tech_2G()
        if y == 'GSM':
            if x == 0:
                pass
            elif x == 1:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
            elif x == 2:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[4] = True
                input_list_bool[5] = True
                input_list_bool[6] = True
                input_list_bool[7] = True
            elif x == 3:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[4] = True
                input_list_bool[5] = True
                input_list_bool[6] = True
                input_list_bool[7] = True
                input_list_bool[8] = True
                input_list_bool[9] = True
                input_list_bool[10] = True
                input_list_bool[11] = True
            elif x == 4:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[4] = True
                input_list_bool[5] = True
                input_list_bool[6] = True
                input_list_bool[7] = True
                input_list_bool[8] = True
                input_list_bool[9] = True
                input_list_bool[10] = True
                input_list_bool[11] = True
                input_list_bool[12] = True
                input_list_bool[13] = True
                input_list_bool[14] = True
                input_list_bool[15] = True
        elif y == 'GSM+DCS':
            if x == 0:
                pass
            elif x == 1:
                input_list_bool[0] = True
                self.bcf_gS1.setText(self.get_bcfID())
                input_list_bool[1] = True
                self.bts_gS1.setText(self.get_bcfID())
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[16] = True
                input_list_bool[17] = True
                input_list_bool[18] = True
                input_list_bool[19] = True
            elif x == 2:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[16] = True
                input_list_bool[17] = True
                input_list_bool[18] = True
                input_list_bool[19] = True
                input_list_bool[4] = True
                input_list_bool[5] = True
                input_list_bool[6] = True
                input_list_bool[7] = True
                input_list_bool[20] = True
                input_list_bool[21] = True
                input_list_bool[22] = True
                input_list_bool[23] = True
            elif x == 3:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[16] = True
                input_list_bool[17] = True
                input_list_bool[18] = True
                input_list_bool[19] = True
                input_list_bool[4] = True
                input_list_bool[5] = True
                input_list_bool[6] = True
                input_list_bool[7] = True
                input_list_bool[20] = True
                input_list_bool[21] = True
                input_list_bool[22] = True
                input_list_bool[23] = True
                input_list_bool[8] = True
                input_list_bool[9] = True
                input_list_bool[10] = True
                input_list_bool[11] = True
                input_list_bool[24] = True
                input_list_bool[25] = True
                input_list_bool[26] = True
                input_list_bool[27] = True
            elif x == 4:
                input_list_bool[0] = True
                input_list_bool[1] = True
                input_list_bool[2] = True
                input_list_bool[3] = True
                input_list_bool[16] = True
                input_list_bool[17] = True
                input_list_bool[18] = True
                input_list_bool[19] = True
                input_list_bool[4] = True
                input_list_bool[5] = True
                input_list_bool[6] = True
                input_list_bool[7] = True
                input_list_bool[20] = True
                input_list_bool[21] = True
                input_list_bool[22] = True
                input_list_bool[23] = True
                input_list_bool[8] = True
                input_list_bool[9] = True
                input_list_bool[10] = True
                input_list_bool[11] = True
                input_list_bool[24] = True
                input_list_bool[25] = True
                input_list_bool[26] = True
                input_list_bool[27] = True
                input_list_bool[12] = True
                input_list_bool[13] = True
                input_list_bool[14] = True
                input_list_bool[15] = True
                input_list_bool[28] = True
                input_list_bool[29] = True
                input_list_bool[30] = True
                input_list_bool[31] = True
        input_list = [self.bcf_gS1.setEnabled(input_list_bool[0]), self.bts_gS1.setEnabled(input_list_bool[1]),
                      self.gS1_CI.setEnabled(input_list_bool[2]), self.gS1_TRX.setEnabled(input_list_bool[3]),
                      self.bcf_gS2.setEnabled(input_list_bool[4]), self.bts_gS2.setEnabled(input_list_bool[5]),
                      self.gS2_CI.setEnabled(input_list_bool[6]), self.gS2_TRX.setEnabled(input_list_bool[7]),
                      self.bcf_gS3.setEnabled(input_list_bool[8]), self.bts_gS3.setEnabled(input_list_bool[9]),
                      self.gS3_CI.setEnabled(input_list_bool[10]), self.gS3_TRX.setEnabled(input_list_bool[11]),
                      self.bcf_gS4.setEnabled(input_list_bool[12]), self.bts_gS4.setEnabled(input_list_bool[13]),
                      self.gS4_CI.setEnabled(input_list_bool[14]), self.gS4_TRX.setEnabled(input_list_bool[15]),
                      self.bcf_dS1.setEnabled(input_list_bool[16]), self.bts_dS1.setEnabled(input_list_bool[17]),
                      self.dS1_CI.setEnabled(input_list_bool[18]), self.dS1_TRX.setEnabled(input_list_bool[19]),
                      self.bcf_dS2.setEnabled(input_list_bool[20]), self.bts_dS2.setEnabled(input_list_bool[21]),
                      self.dS2_CI.setEnabled(input_list_bool[22]), self.dS2_TRX.setEnabled(input_list_bool[23]),
                      self.bcf_dS3.setEnabled(input_list_bool[24]), self.bts_dS3.setEnabled(input_list_bool[25]),
                      self.dS3_CI.setEnabled(input_list_bool[26]), self.dS3_TRX.setEnabled(input_list_bool[27]),
                      self.bcf_dS4.setEnabled(input_list_bool[28]), self.bts_dS4.setEnabled(input_list_bool[29]),
                      self.dS4_CI.setEnabled(input_list_bool[30]), self.dS4_TRX.setEnabled(input_list_bool[31])]
        return input_list

    def CI_TRX_change(self):
        input_list_values = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
        x = self.get_num_sectors()
        y = self.get_tech_2G()
        if y == 'GSM':
            if x == 0:
                pass
            elif x == 1:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()

            elif x == 2:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[2] = self.gS2_CI.text()
                input_list_values[3] = self.gS2_TRX.text()

            elif x == 3:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[2] = self.gS2_CI.text()
                input_list_values[3] = self.gS2_TRX.text()
                input_list_values[4] = self.gS3_CI.text()
                input_list_values[5] = self.gS3_TRX.text()

            elif x == 4:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[2] = self.gS2_CI.text()
                input_list_values[3] = self.gS2_TRX.text()
                input_list_values[4] = self.gS3_CI.text()
                input_list_values[5] = self.gS3_TRX.text()
                input_list_values[6] = self.gS4_CI.text()
                input_list_values[7] = self.gS4_TRX.text()
        elif y == 'GSM+DCS':
            if x == 0:
                pass
            elif x == 1:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[8] = self.dS1_CI.text()
                input_list_values[9] = self.dS1_TRX.text()
            elif x == 2:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[2] = self.gS2_CI.text()
                input_list_values[3] = self.gS2_TRX.text()
                input_list_values[8] = self.dS1_CI.text()
                input_list_values[9] = self.dS1_TRX.text()
                input_list_values[10] = self.dS2_CI.text()
                input_list_values[11] = self.dS2_TRX.text()

            elif x == 3:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[2] = self.gS2_CI.text()
                input_list_values[3] = self.gS2_TRX.text()
                input_list_values[4] = self.gS3_CI.text()
                input_list_values[5] = self.gS3_TRX.text()
                input_list_values[8] = self.dS1_CI.text()
                input_list_values[9] = self.dS1_TRX.text()
                input_list_values[10] = self.dS2_CI.text()
                input_list_values[11] = self.dS2_TRX.text()
                input_list_values[12] = self.dS3_CI.text()
                input_list_values[13] = self.dS3_TRX.text()

            elif x == 4:
                input_list_values[0] = self.gS1_CI.text()
                input_list_values[1] = self.gS1_TRX.text()
                input_list_values[2] = self.gS2_CI.text()
                input_list_values[3] = self.gS2_TRX.text()
                input_list_values[4] = self.gS3_CI.text()
                input_list_values[5] = self.gS3_TRX.text()
                input_list_values[6] = self.gS4_CI.text()
                input_list_values[7] = self.gS4_TRX.text()
                input_list_values[8] = self.dS1_CI.text()
                input_list_values[9] = self.dS1_TRX.text()
                input_list_values[10] = self.dS2_CI.text()
                input_list_values[11] = self.dS2_TRX.text()
                input_list_values[12] = self.dS3_CI.text()
                input_list_values[13] = self.dS3_TRX.text()
                input_list_values[14] = self.dS4_CI.text()
                input_list_values[15] = self.dS4_TRX.text()

        input_list = [self.gS1_CI.setText(input_list_values[0]), self.gS1_TRX.setText(input_list_values[1]),
                      self.gS2_CI.setText(input_list_values[2]), self.gS2_TRX.setText(input_list_values[3]),
                      self.gS3_CI.setText(input_list_values[4]), self.gS3_TRX.setText(input_list_values[5]),
                      self.gS4_CI.setText(input_list_values[6]), self.gS4_TRX.setText(input_list_values[7]),
                      self.dS1_CI.setText(input_list_values[8]), self.dS1_TRX.setText(input_list_values[9]),
                      self.dS2_CI.setText(input_list_values[10]), self.dS2_TRX.setText(input_list_values[11]),
                      self.dS3_CI.setText(input_list_values[12]), self.dS3_TRX.setText(input_list_values[13]),
                      self.dS4_CI.setText(input_list_values[14]), self.dS4_TRX.setText(input_list_values[15])]
        return input_list

    def get_siteNameG(self):
        siteName_G = self.get_bcfID() + "_O_" + self.get_siteName() + "_" + self.get_siteCode()
        print(siteName_G)
        return siteName_G

    def get_siteNameU(self):
        siteName_U = "U_O_" + self.get_siteName() + "_" + self.get_siteCode()
        print(siteName_U)
        return siteName_U

    def get_siteNameL(self):
        siteName_L = "L_O_" + self.get_siteName() + "_" + self.get_siteCode()
        print(siteName_L)
        return siteName_L

    def set_2G_config(self):
        config_array_value = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
        x = self.get_num_sectors()
        y = self.get_tech_2G()
        if y == 'GSM':
            if x == 1:
                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
            elif x == 2:
                bts2 = int(self.get_bcfID()) + 1
                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[2] = self.get_bcfID()
                config_array_value[3] = str(bts2)
            elif x == 3:
                bts2 = int(self.get_bcfID()) + 1
                bts3 = int(self.get_bcfID()) + 2
                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[2] = self.get_bcfID()
                config_array_value[3] = str(bts2)
                config_array_value[4] = self.get_bcfID()
                config_array_value[5] = str(bts3)
            elif x == 4:
                bts2 = int(self.get_bcfID()) + 1
                bts3 = int(self.get_bcfID()) + 2
                bts4 = int(self.get_bcfID()) + 3
                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[2] = self.get_bcfID()
                config_array_value[3] = str(bts2)
                config_array_value[4] = self.get_bcfID()
                config_array_value[5] = str(bts3)
                config_array_value[6] = self.get_bcfID()
                config_array_value[7] = str(bts4)

        elif y == 'GSM+DCS':
            if x == 1:
                bts5 = int(self.get_bcfID()) + 4
                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[8] = self.get_bcfID()
                config_array_value[9] = str(bts5)
            elif x == 2:
                bts2 = int(self.get_bcfID()) + 1
                bts5 = int(self.get_bcfID()) + 4
                bts6 = int(self.get_bcfID()) + 5

                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[2] = self.get_bcfID()
                config_array_value[3] = str(bts2)
                config_array_value[8] = self.get_bcfID()
                config_array_value[9] = str(bts5)
                config_array_value[10] = self.get_bcfID()
                config_array_value[11] = str(bts6)
            elif x == 3:
                bts2 = int(self.get_bcfID()) + 1
                bts3 = int(self.get_bcfID()) + 2
                bts5 = int(self.get_bcfID()) + 4
                bts6 = int(self.get_bcfID()) + 5
                bts7 = int(self.get_bcfID()) + 6

                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[2] = self.get_bcfID()
                config_array_value[3] = str(bts2)
                config_array_value[4] = self.get_bcfID()
                config_array_value[5] = str(bts3)
                config_array_value[8] = self.get_bcfID()
                config_array_value[9] = str(bts5)
                config_array_value[10] = self.get_bcfID()
                config_array_value[11] = str(bts6)
                config_array_value[12] = self.get_bcfID()
                config_array_value[13] = str(bts7)

            elif x == 4:
                bts2 = int(self.get_bcfID()) + 1
                bts3 = int(self.get_bcfID()) + 2
                bts4 = int(self.get_bcfID()) + 3
                bts5 = int(self.get_bcfID()) + 4
                bts6 = int(self.get_bcfID()) + 5
                bts7 = int(self.get_bcfID()) + 6
                bts8 = int(self.get_bcfID()) + 7

                config_array_value[0] = self.get_bcfID()
                config_array_value[1] = self.get_bcfID()
                config_array_value[2] = self.get_bcfID()
                config_array_value[3] = str(bts2)
                config_array_value[4] = self.get_bcfID()
                config_array_value[5] = str(bts3)
                config_array_value[6] = self.get_bcfID()
                config_array_value[7] = str(bts4)
                config_array_value[8] = self.get_bcfID()
                config_array_value[9] = str(bts5)
                config_array_value[10] = self.get_bcfID()
                config_array_value[11] = str(bts6)
                config_array_value[12] = self.get_bcfID()
                config_array_value[13] = str(bts7)
                config_array_value[14] = self.get_bcfID()
                config_array_value[15] = str(bts8)

        config_array = [self.bcf_gS1.setText(config_array_value[0]), self.bts_gS1.setText(config_array_value[1]),
                        self.bcf_gS2.setText(config_array_value[2]), self.bts_gS2.setText(config_array_value[3]),
                        self.bcf_gS3.setText(config_array_value[4]), self.bts_gS3.setText(config_array_value[5]),
                        self.bcf_gS4.setText(config_array_value[6]), self.bts_gS4.setText(config_array_value[7]),
                        self.bcf_dS1.setText(config_array_value[8]), self.bts_dS1.setText(config_array_value[9]),
                        self.bcf_dS2.setText(config_array_value[10]), self.bts_dS2.setText(config_array_value[11]),
                        self.bcf_dS3.setText(config_array_value[12]), self.bts_dS3.setText(config_array_value[13]),
                        self.bcf_dS4.setText(config_array_value[14]), self.bts_dS4.setText(config_array_value[15])]
        return config_array

    def run_pABIS_macro(self):
        fileName = QFileDialog.getOpenFileName(filter="Excel (*.xlsm)", caption="Please choose PABIS Macro tool")
        if fileName[1] == "":
            QMessageBox.about(self, 'Caution', "PABIS Macro not uploaded")
        else:
            path = fileName[0]
            pABIS_tool = xw.Book(path)
            QMessageBox.about(self, 'Done', "PABIS Macro uploaded")
            sheet = pABIS_tool.sheets[0]
            sheet.range("E2:J9").clear_contents()
            sheet.range('A2').value = self.get_bcfID()
            sheet.range('B2').value = self.get_siteNameG()
            sheet.range('L2').value = self.get_bscName()
            sheet.range('M2').value = self.get_bscID()
            sheet.range('E2').value = self.get_bcfID()

            bcf_list = [self.bcf_gS1.text(), self.bcf_gS2.text(), self.bcf_gS3.text(), self.bcf_gS4.text(),
                        self.bcf_dS1.text(), self.bcf_dS2.text(), self.bcf_dS3.text(), self.bcf_dS4.text()]
            bts_list = [self.bts_gS1.text(), self.bts_gS2.text(), self.bts_gS3.text(), self.bts_gS4.text(),
                        self.bts_dS1.text(), self.bts_dS2.text(), self.bts_dS3.text(), self.bts_dS4.text()]
            cellID_list = [self.gS1_CI.text(), self.gS2_CI.text(), self.gS3_CI.text(), self.gS4_CI.text(),
                           self.dS1_CI.text(), self.dS2_CI.text(), self.dS3_CI.text(), self.dS4_CI.text()]
            config_list = [self.gS1_TRX.text(), self.gS2_TRX.text(), self.gS3_TRX.text(), self.gS4_TRX.text(),
                           self.dS1_TRX.text(), self.dS2_TRX.text(), self.dS3_TRX.text(), self.dS4_TRX.text()]

            df_bcf = pd.DataFrame(bcf_list)
            df_bcf = df_bcf.loc[~(df_bcf == '').all(axis=1)]
            df_bts = pd.DataFrame(bts_list)
            df_bts = df_bts.loc[~(df_bts == '').all(axis=1)]
            df_CI = pd.DataFrame(cellID_list)
            df_CI = df_CI.loc[~(df_CI == '').all(axis=1)]
            df_trx = pd.DataFrame(config_list)
            df_trx = df_trx.loc[~(df_trx == '').all(axis=1)]
            lac_list = [self.get_lacG()] * len(df_bcf.index)
            rac_list = [self.get_racG()] * len(df_bcf.index)
            df_lacG = pd.DataFrame(lac_list)
            df_racG = pd.DataFrame(rac_list)

            sheet['E2'].options(pd.DataFrame, header=0, index=False, expand='table').value = df_bcf
            sheet['F2'].options(pd.DataFrame, header=0, index=False, expand='table').value = df_bts
            sheet['G2'].options(pd.DataFrame, header=0, index=False, expand='table').value = df_CI
            sheet['H2'].options(pd.DataFrame, header=0, index=False, expand='table').value = df_trx
            sheet['I2'].options(pd.DataFrame, header=0, index=False, expand='table').value = df_lacG
            sheet['J2'].options(pd.DataFrame, header=0, index=False, expand='table').value = df_racG

            macro1 = pABIS_tool.macro("legacyCIQ")
            macro1()
            pABIS_tool.save()
            pABIS_tool.close()
            return True

    def create_2G_dataframe(self):
        x = self.run_pABIS_macro()
        while x:
            x = False
            fileName = QFileDialog.getOpenFileName(filter="Excel (*.xlsx *.xls)",
                                                   caption="Select Macro output that was just Saved!")
            if fileName[1] == "":
                QMessageBox.about(self, 'Caution', "No file uploaded")
            else:
                path = fileName[0]
                cwo_2G_final = load_workbook(path)
                bcf_sheet = cwo_2G_final.worksheets[3]
                bcf_sheet['D2'].value = self.get_cuPlaneIP()
                bcf_sheet['E2'].value = self.get_mngtIP()
                bcf_sheet['H2'].value = self.get_usedETP()

                lapd_sheet = cwo_2G_final.worksheets[2]
                lapd_sheet_lastrow = lapd_sheet.max_row

                for row in lapd_sheet.iter_rows(min_col=7, min_row=2, max_col=7, max_row=lapd_sheet_lastrow):
                    for cell in row:
                        cell.value = self.get_bcsuID()

                response = QFileDialog.getSaveFileName(caption='Save your CWO', directory=f'{self.siteName_SRAN.text()}2G CWO',
                                                       filter="Excel (*.xlsx *.xls *.csv)")
                if response[0] != "":
                    cwo_2G_final.save(response[0])
                    QMessageBox.about(self, 'Saved', f" {self.get_siteNameG()}2G CWO saved Successfully")
                else:
                    QMessageBox.about(self, 'Saved', "File Not Saved!")

                # os.remove("2G_CWO.xlsx")

    def create_3G_dataframe(self):
        wb_3G = Workbook()

        wb_3G.create_sheet('Cells Data')

        self.cellU_df[self.lacU_column.currentText()] = self.cellU_df[
            self.lacU_column.currentText()].apply(lambda x: '{0:0>5}'.format(x))
        self.cellU_df[self.cellID_column.currentText()] = self.cellU_df[
            self.cellID_column.currentText()].apply(lambda x: '{0:0>5}'.format(x))

        self.cellU_df['LOCN'] = "2012" + self.cellU_df[self.cz_column.currentText()].astype(str) + "0" + \
                                self.cellU_df[self.lacU_column.currentText()].astype(str) + self.cellU_df[
                                    self.cellID_column.currentText()].astype(str)

        wb_3G.create_sheet('Site Data')
        data_3G_df = pd.DataFrame(
            {'RNC': [self.get_rncID()], 'WBTS': [self.get_wbtsID()], 'Site Name': [self.get_siteNameU()],
             'Site Code': [self.get_siteCode()], 'Management IP': [self.get_mngtIP()],
             'ICSU': [self.get_rncCard()], 'SCTP port': [self.get_portSCTP()]})

        wb_3G.create_sheet('IP Designs')
        empty_df = pd.DataFrame()

        wb_3G.create_sheet('BW')

        bw_3G_df = pd.DataFrame({'Site Name': [self.get_siteNameU()], 'Site Code': [self.get_siteCode()],
                                 'Voice Committed BW': [self.get_voiceCommittedBW()],
                                 'Voice Route BW': [self.get_voiceRouteBW()],
                                 'Default Committed BW': [self.get_defaultCommittedBW()],
                                 'Default Route BW': [self.get_defaultRouteBW()], 'Route BW': [self.get_routeBW()]})

        response = QFileDialog.getSaveFileName(caption='Save your 3G CWO',
                                               directory=f'{self.get_siteNameU()} 3G CWO',
                                               filter="Excel (*.xlsx *.xls *.csv)")
        if response[0] == "":
            QMessageBox.about(self, 'Caution', "Please specify save location!")
        else:
            writer = pd.ExcelWriter(response[0], engine='openpyxl')
            self.cellU_df.to_excel(writer, sheet_name='Cells Data', index=False)
            empty_df.to_excel(writer, sheet_name='IP Designs', index=False)
            data_3G_df.to_excel(writer, sheet_name='Data', index=False)
            bw_3G_df.to_excel(writer, sheet_name='BW', index=False)

            writer.save()
            QMessageBox.about(self, 'Done', f"{self.get_siteNameU()} 3G CWO saved successfully!")

    def create_4G_dataframe(self):
        fileName = QFileDialog.getOpenFileName(filter="Excel (*.xlsx *.xls *.csv)", caption="Please choose LTE CWO template")
        if fileName[1] == "":
            QMessageBox.about(self, 'Caution', "LTE CWO template not uploaded")
        else:
            path = fileName[0]
            template_CWO_LTE = load_workbook(path)
            template_CWO_LTE.save('tempLTE.xlsx')
            QMessageBox.about(self, 'Done', "LTE CWO template uploaded")

            core_df = self.cellL_df[[self.eNodeB_LTE.currentText(), self.localCI_LTE.currentText(),
                                     self.cellName_LTE.currentText(), self.cellID_LTE.currentText()]].copy()

            core_df.insert(0, 'Site Code', self.get_siteCode(), allow_duplicates=False)
            core_df.insert(2, 'LTE NE Name', self.get_siteNameL(), allow_duplicates=False)

            core_df[self.cellID_LTE.currentText()] = core_df[
                self.cellID_LTE.currentText()].apply(lambda x: '{0:0>7}'.format(x))

            core_df['LOCN'] = '2012' + self.cellL_df[self.cz_LTE_column.currentText()].astype(str) + '0012' + \
                              core_df[self.cellID_LTE.currentText()].astype(str)

            writer = pd.ExcelWriter('tempLTE.xlsx', engine='openpyxl', mode="a", if_sheet_exists="replace")
            self.cellL_df.to_excel(writer, sheet_name='Radio Data_1800', index=False)
            core_df.to_excel(writer, sheet_name='Core data', index=False)
            writer.save()
            writer.close()

            lte_template = load_workbook('tempLTE.xlsx')
            secGW_sheet = lte_template.worksheets[8]
            secGW_sheet['A2'].value = self.get_siteNameL()
            secGW_sheet['B2'].value = self.get_siteCode()
            secGW_sheet['C2'].value = self.get_secIP()
            secGW_sheet['D2'].value = self.get_secGW()
            secGW_sheet['E2'].value = self.get_coreSite()

            netAct_sheet = lte_template.worksheets[3]
            netAct_sheet['A3'].value = self.get_siteNameL()
            netAct_sheet['B3'].value = self.get_siteCode()
            netAct_sheet['H3'].value = self.get_mngtGW()

            if self.get_site_area() == "AL":
                del lte_template['MME Delta']
                del lte_template['MME Upper']
            elif self.get_site_area() == "DE":
                del lte_template['MME Alex']
                del lte_template['MME Upper']

            elif self.get_site_area() == "UP" or self.get_site_area() == "SI":
                del lte_template['MME Delta']
                del lte_template['MME Alex']

            response = QFileDialog.getSaveFileName(
                caption='Save your 4G CWO', directory=f'{self.get_siteNameU()} 4G CWO',
                filter="Excel (*.xlsx *.xls *.csv)")

            if response[0] != "":
                lte_template.save(response[0])
                QMessageBox.about(self, 'Saved', f" {self.get_siteNameL()} 4G CWO saved Successfully")

                # wb = load_workbook(response[0])
                # sheet_name = 'Radio Data_1800'
                # idx = wb.sheetnames.index(sheet_name)
                # ws = wb[sheet_name]
                # wb.remove(ws)
                # wb.create_sheet(sheet_name, idx)
                #
                # sheet_name_1 = 'Core data'
                # idx_1 = wb.sheetnames.index(sheet_name_1)
                # ws_1 = wb[sheet_name_1]
                # wb.remove(ws_1)
                # wb.create_sheet(sheet_name_1, idx_1)
                # wb.save(response[0])
            else:
                QMessageBox.about(self, 'Caution', "File Not Saved!")

    def validate_inputs(self):
        self.cwo2G.setEnabled(False)
        self.cwo3G.setEnabled(False)
        self.cwo4G.setEnabled(False)
        validation_list = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1,
                           1, 1, 1, 1]

        if self.siteName_SRAN.text() == '':
            self.siteName_SRAN.setStyleSheet("background-color: pink;")
            validation_list[0] = 1
        else:
            self.siteName_SRAN.setStyleSheet("background-color: lightgreen;")
            validation_list[0] = 0

        if self.get_site_area() == "AL":
            self.area_label.setStyleSheet("background-color: lightgreen")
            self.area_label.setText('Alex')
            validation_list[1] = 0

        elif self.get_site_area() == "DE":
            self.area_label.setStyleSheet("background-color: lightgreen")
            self.area_label.setText('Delta')
            validation_list[1] = 0

        elif self.get_site_area() == "UP" or self.get_site_area() == "SI":
            self.area_label.setStyleSheet("background-color: lightgreen")
            self.area_label.setText('SI/UP')
            validation_list[1] = 0

        else:
            self.area_label.setStyleSheet("background-color: red")
            self.area_label.setText('INVALID!')
            validation_list[1] = 1

        if self.mngtIP.text() != "":
            self.mngtIP.setStyleSheet("background-color: lightgreen;")
            validation_list[2] = 0

        else:
            self.mngtIP.setStyleSheet("background-color: pink;")
            validation_list[2] = 1

        if self.mngtGW.text() != "":
            self.mngtGW.setStyleSheet("background-color: lightgreen;")
            validation_list[3] = 0

        else:
            self.mngtGW.setStyleSheet("background-color: pink;")
            validation_list[3] = 1

        if self.rncID.text() != "":
            self.rncID.setStyleSheet("background-color: lightgreen;")
            validation_list[4] = 0

        else:
            self.rncID.setStyleSheet("background-color: pink;")
            validation_list[4] = 1

        if self.wbtsID.text() != "":
            self.wbtsID.setStyleSheet("background-color: lightgreen;")
            validation_list[5] = 0

        else:
            self.wbtsID.setStyleSheet("background-color: pink;")
            validation_list[5] = 1

        if self.portSCTP.text() != "":
            self.portSCTP.setStyleSheet("background-color: lightgreen;")
            validation_list[6] = 0

        else:
            self.portSCTP.setStyleSheet("background-color: pink;")
            validation_list[6] = 1

        if self.rncCard.text() != "":
            self.rncCard.setStyleSheet("background-color: lightgreen;")
            validation_list[7] = 0

        else:
            self.rncCard.setStyleSheet("background-color: pink;")
            validation_list[7] = 1

        if self.secIP.text() != "":
            self.secIP.setStyleSheet("background-color: lightgreen;")
            validation_list[8] = 0

        else:
            self.secIP.setStyleSheet("background-color: pink;")
            validation_list[8] = 1

        if self.secGW.text() != "":
            self.secGW.setStyleSheet("background-color: lightgreen;")
            validation_list[9] = 0

        else:
            self.secGW.setStyleSheet("background-color: pink;")
            validation_list[9] = 1

        if self.coreSite.text() != "":
            self.coreSite.setStyleSheet("background-color: lightgreen;")
            validation_list[10] = 0

        else:
            self.coreSite.setStyleSheet("background-color: pink;")
            validation_list[10] = 1

        if self.serverCA.text() != "":
            self.serverCA.setStyleSheet("background-color: lightgreen;")
            validation_list[11] = 0

        else:
            self.serverCA.setStyleSheet("background-color: pink;")
            validation_list[11] = 1

        if self.bcfID.text() != '0':
            self.bcfID.setStyleSheet("background-color: lightgreen;")
            validation_list[12] = 0

        else:
            self.bcfID.setStyleSheet("background-color: pink;")
            validation_list[12] = 1

        if self.bscID.text() != '':
            self.bscID.setStyleSheet("background-color: lightgreen;")
            validation_list[13] = 0

        else:
            self.bscID.setStyleSheet("background-color: pink;")
            validation_list[13] = 1

        if self.bcsuID.text() != '':
            self.bcsuID.setStyleSheet("background-color: lightgreen;")
            validation_list[14] = 0

        else:
            self.bcsuID.setStyleSheet("background-color: pink;")
            validation_list[14] = 1

        if self.lacGSM.text() != '0':
            self.lacGSM.setStyleSheet("background-color: lightgreen;")
            validation_list[15] = 0

        else:
            self.lacGSM.setStyleSheet("background-color: pink;")
            validation_list[15] = 1

        if self.racG.text() != '0':
            self.racG.setStyleSheet("background-color: lightgreen;")
            validation_list[16] = 0

        else:
            self.racG.setStyleSheet("background-color: pink;")
            validation_list[16] = 1

        if self.usedETP.text() != '':
            self.usedETP.setStyleSheet("background-color: lightgreen;")
            validation_list[17] = 0

        else:
            self.usedETP.setStyleSheet("background-color: pink;")
            validation_list[17] = 1

        if self.cuIP.text() != '':
            self.cuIP.setStyleSheet("background-color: lightgreen;")
            validation_list[18] = 0

        else:
            self.cuIP.setStyleSheet("background-color: pink;")
            validation_list[18] = 1

        if self.bcf_gS1.text() != '':
            if self.gS1_CI.text() == '':
                self.gS1_CI.setStyleSheet("background-color: pink;")
                validation_list[19] = 1
            else:
                self.gS1_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[19] = 0

            if self.gS1_TRX.text() == '':
                self.gS1_TRX.setStyleSheet("background-color: pink;")
                validation_list[20] = 1

            else:
                self.gS1_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[20] = 0

        if self.bcf_gS2.text() != '':
            if self.gS2_CI.text() == '':
                self.gS2_CI.setStyleSheet("background-color: pink;")
                validation_list[21] = 1

            else:
                self.gS2_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[21] = 0

            if self.gS2_TRX.text() == '':
                self.gS2_TRX.setStyleSheet("background-color: pink;")
                validation_list[22] = 1

            else:
                self.gS2_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[22] = 0

        if self.bcf_gS3.text() != '':
            if self.gS3_CI.text() == '':
                self.gS3_CI.setStyleSheet("background-color: pink;")
                validation_list[23] = 1

            else:
                self.gS3_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[23] = 0

            if self.gS3_TRX.text() == '':
                self.gS3_TRX.setStyleSheet("background-color: pink;")
                validation_list[24] = 1

            else:
                self.gS3_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[24] = 0

        if self.bcf_gS4.text() != '':
            if self.gS4_CI.text() == '':
                self.gS4_CI.setStyleSheet("background-color: pink;")
                validation_list[25] = 1

            else:
                self.gS4_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[25] = 0

            if self.gS4_TRX.text() == '':
                self.gS4_TRX.setStyleSheet("background-color: pink;")
                validation_list[26] = 1

            else:
                self.gS4_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[26] = 0

        if self.bcf_dS1.text() != '':
            if self.dS1_CI.text() == '':
                self.dS1_CI.setStyleSheet("background-color: pink;")
                validation_list[27] = 1

            else:
                self.dS1_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[27] = 0

            if self.dS1_TRX.text() == '':
                self.dS1_TRX.setStyleSheet("background-color: pink;")
                validation_list[28] = 1

            else:
                self.dS1_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[28] = 0

        if self.bcf_dS2.text() != '':
            if self.dS2_CI.text() == '':
                self.dS2_CI.setStyleSheet("background-color: pink;")
                validation_list[29] = 1

            else:
                self.dS2_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[29] = 0

            if self.dS2_TRX.text() == '':
                self.dS2_TRX.setStyleSheet("background-color: pink;")
                validation_list[30] = 1

            else:
                self.dS2_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[30] = 0

        if self.bcf_dS3.text() != '':
            if self.dS3_CI.text() == '':
                self.dS3_CI.setStyleSheet("background-color: pink;")
                validation_list[31] = 1

            else:
                self.dS3_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[31] = 0

            if self.dS3_TRX.text() == '':
                self.dS3_TRX.setStyleSheet("background-color: pink;")
                validation_list[32] = 1

            else:
                self.dS3_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[32] = 0

        if self.bcf_dS4.text() != '':
            if self.dS4_CI.text() == '':
                self.dS4_CI.setStyleSheet("background-color: pink;")
                validation_list[33] = 1

            else:
                self.dS4_CI.setStyleSheet("background-color: lightgreen;")
                validation_list[33] = 0

            if self.dS4_TRX.text() == '':
                self.dS4_TRX.setStyleSheet("background-color: pink;")
                validation_list[34] = 1

            else:
                self.dS4_TRX.setStyleSheet("background-color: lightgreen;")
                validation_list[34] = 0

        if not self.bcf_gS1.isEnabled():
            self.gS1_CI.setStyleSheet("background-color: #ececec;")
            self.gS1_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_gS2.isEnabled():
            self.gS2_CI.setStyleSheet("background-color: #ececec;")
            self.gS2_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_gS3.isEnabled():
            self.gS3_CI.setStyleSheet("background-color: #ececec;")
            self.gS3_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_gS4.isEnabled():
            self.gS4_CI.setStyleSheet("background-color: #ececec;")
            self.gS4_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_dS1.isEnabled():
            self.dS1_CI.setStyleSheet("background-color: #ececec;")
            self.dS1_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_dS2.isEnabled():
            self.dS2_CI.setStyleSheet("background-color: #ececec;")
            self.dS2_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_gS3.isEnabled():
            self.dS3_CI.setStyleSheet("background-color: #ececec;")
            self.dS3_TRX.setStyleSheet("background-color: #ececec;")

        if not self.bcf_dS4.isEnabled():
            self.dS4_CI.setStyleSheet("background-color: #ececec;")
            self.dS4_TRX.setStyleSheet("background-color: #ececec;")

        if sum(validation_list[0:4]) == 0:
            if sum(validation_list[4:8]) == 0:
                self.cwo3G.setEnabled(True)
            else:
                self.cwo3G.setEnabled(False)

            if sum(validation_list[8:12]) == 0:
                self.cwo4G.setEnabled(True)
            else:
                self.cwo4G.setEnabled(False)

            if sum(validation_list[12:19]) == 0:
                if self.get_tech_2G() == 'GSM' and self.get_num_sectors() == 1:
                    if sum(validation_list[19:21]) == 0:
                        self.cwo2G.setEnabled(True)

                elif self.get_tech_2G() == 'GSM' and self.get_num_sectors() == 2:
                    if sum(validation_list[19:23]) == 0:
                        self.cwo2G.setEnabled(True)

                elif self.get_tech_2G() == 'GSM' and self.get_num_sectors() == 3:
                    if sum(validation_list[19:25]) == 0:
                        self.cwo2G.setEnabled(True)

                elif self.get_tech_2G() == 'GSM' and self.get_num_sectors() == 4:
                    if sum(validation_list[19:27]) == 0:
                        self.cwo2G.setEnabled(True)

                elif self.get_tech_2G() == 'GSM+DCS' and self.get_num_sectors() == 1:
                    if sum(validation_list[19:21]) == 0 and sum(validation_list[27:29]) == 0:
                        self.cwo2G.setEnabled(True)
                    else:
                        QMessageBox.about(self, 'Error', "Missing Cell ID / TRX Config inputs!")

                elif self.get_tech_2G() == 'GSM+DCS' and self.get_num_sectors() == 2:
                    if sum(validation_list[19:23]) == 0 and sum(validation_list[27:31]) == 0:
                        self.cwo2G.setEnabled(True)

                elif self.get_tech_2G() == 'GSM+DCS' and self.get_num_sectors() == 3:
                    if sum(validation_list[19:25]) == 0 and sum(validation_list[27:33]) == 0:
                        self.cwo2G.setEnabled(True)

                elif self.get_tech_2G() == 'GSM+DCS' and self.get_num_sectors() == 4:
                    if sum(validation_list[19:27]) == 0 and sum(validation_list[27:35]) == 0:
                        self.cwo2G.setEnabled(True)

            else:
                self.cwo2G.setEnabled(False)


class Manager:
    def __init__(self):
        # Creating App Window
        self.appWindow = appWindow()

        # Start the program
        self.appWindow.show()


#####################
#        MAIN       #
#####################
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    manager = Manager()
    sys.exit(app.exec_())
